from openpyxl.styles import Border, Side,PatternFill,Font,Alignment

def add_borders_top(sheet, start_row=1, end_row=9, columns=['A','B','C','D','E','F','G','H']):
    """
    Add black borders to specified columns and rows in the given worksheet.

    Parameters:
    - sheet: openpyxl.worksheet.worksheet.Worksheet, the worksheet where borders will be added.
    - start_row: int, the starting row number for adding borders.
    - end_row: int, the ending row number for adding borders.
    - columns: list of str, column letters where borders will be applied.
    """
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    for row in range(start_row, end_row + 1):
        for col in columns:
            cell = sheet[f'{col}{row}']
            cell.border = border

# Example usage
def add_borders(sheet, start_row, end_row, columns):
    """
    Add custom borders to specified columns and rows in the given worksheet.
    
    Parameters:
    - sheet: openpyxl.worksheet.worksheet.Worksheet, the worksheet where borders will be added.
    - start_row: int, the starting row number for adding borders.
    - end_row: int, the ending row number for adding borders.
    - columns: list of str, column letters where borders will be applied.
    """
    # Define border styles
    top_border = Border(top=Side(border_style="thin", color="000000"),left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"))
    bottom_border = Border(bottom=Side(border_style="thin", color="000000"),left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"))
    no_border = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"))

    for row in range(start_row, end_row + 1):
        # Determine the border style based on the row number
        if row % 3 == 1:  # e.g., Row 10, 13, 16, etc. (1, 4, 7, ...)
            border_style = top_border
        elif row % 3 == 2:  # e.g., Row 11, 14, 17, etc. (2, 5, 8, ...)
            border_style = no_border
        elif row % 3 == 0:  # e.g., Row 12, 15, 18, etc. (3, 6, 9, ...)
            border_style = bottom_border

        # Apply the selected border style to each cell in the specified columns
        for col in columns:
            cell = sheet[f'{col}{row}']
            cell.border = border_style


def adjust_column_widths(sheet, column_widths):
  for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
#Formula Adder

def addcolor(sheet):
        sheet['C5'].fill = PatternFill(start_color='87CEFA',end_color='87CEFA',fill_type='solid')
def make_bold(sheet):
    bold_font = Font(bold=True)
    cells=['A4','A5','A6','D1','D4','D6','C5']
    for cell in cells:
        sheet[cell].font=bold_font
def adjust_alignment(sheet):
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment	
   