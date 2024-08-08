from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side,Alignment
from style import add_borders,add_borders_top,adjust_column_widths,addcolor,make_bold,adjust_alignment
from addFormula import add_formula_to_column,add_formula_to_column2,add_formula_to_column3

def generateExcel(sample_data,date):
    file_name = date+'.xlsx' 
    bold_font = Font(bold=True)

    wb = Workbook()
    ws1 = wb.active
    #FIRST SHEET
    ws1.title = 'MOTHER'
    ws1['D1'].font=bold_font
    merge_ranges = ['A1:B3','C1:C3','D1:H3','A4:B4','C4','D4:E4','F4:H4','A5:B5','C5:H5','A6:B7','C6:C7','D6:E7','F6:H7','A8:B9', 'C8:C9','D8:E9','F8:H9','A10:A12','B10:B12','C10:C12','D10:E12','F10:F12','G10:G12','H10:H12']
    static_data=['Type of Vehicle:','MARUTI ECO','PPP ambulance / Empanelled vehicle (strike off which is not applicable)','Name of the Block:','GALSI – II','District:','BURDWAN:','Name of the NGO:','MOTHER','Name of the operator:','TARUN KUMAR CHATTOPADHYAY','Vehicle No.','WB-41J7245','','','YEAR:',date,'Sl. No.','Name of the Beneficiaries','Address','Type of vouchers','Distance travelled (KM.)','Date of travel','Amount claimed (Rs.)']
    i=0
    for cell in merge_ranges:
        if(":" in cell):
            ws1.merge_cells(cell)
            ws1[cell.split(":")[0]]=static_data[i]
        else:
            ws1[cell]=static_data[i]
        i+=1


    n=len(sample_data)

    start=13 #

    distance_per_voucher=[]
    for i in range(n):
        ws1[f'A${start+1}']=i+1
        ws1[f'B${start+1}']=sample_data[i][1].upper()
        ws1[f'C${start+1}']=sample_data[i][2]
        ws1[f'C${start+2}']="Burdwan"
            
        ws1[f'D${start}']="V1"
        ws1[f'D${start+1}']="V2"  
        ws1[f'D${start+2}']="V3"
        if(isinstance(sample_data[i][3], list)):
            for j in range(len(sample_data[i][3])):
                ws1[f'F${start+j}']=sample_data[i][4][j]
                ws1[f'G${start+j}']=sample_data[i][5][j]
                distance_per_voucher.append(sample_data[i][4][j])

                

        else:
            distance_per_voucher.append(sample_data[i][4])
            ws1[f'F${start+sample_data[i][3]-1}']=sample_data[i][4]
            ws1[f'G${start+sample_data[i][3]-1}']=sample_data[i][5]


        start+=3
    #Add font style
    make_bold(ws1)
    #Add color
    addcolor(ws1)
    add_borders(ws1, start_row=10, end_row=ws1.max_row, columns=['A', 'B', 'C'])
    add_borders_top(ws1, 10, ws1.max_row,['D','E','F','G','H'])
    add_borders_top(ws1)
    #Add formula
    add_formula_to_column(ws1,13)
    add_formula_to_column2(ws1,13)
    add_formula_to_column3(ws1,13)
    adjust_column_widths(ws1,{'A':8,'B':30,'C':28,'D':8,'E':15,'F':20,'G':15,'H':20})

    # Set the alignment for each cell in the sheet
    adjust_alignment(ws1)
    #adjust_row_heights(ws1)

    ws2 = wb.create_sheet(title='Sheet2')
    ws2['B4'] = 'Voucher SL No'
    ws2['C4'] = 'KM PER VOUCHER'
    ws2['D4'] = 'BILL AMT PER VOUCHER Rs.'
    v=381
    s=5
    for i in range(len(distance_per_voucher)):
        ws2[f'B${s+i}']=v+i
        ws2[f'C${s+i}']=distance_per_voucher[i]
    adjust_column_widths(ws2,{'B':15,'C':17,'D':26})
    add_formula_to_column2(ws2, 5,'C',4)
    adjust_alignment(ws2)
    add_borders_top(ws2,4,ws2.max_row,['B','C','D'])
    add_formula_to_column3(ws2,5,'D')


    wb.save(file_name)
